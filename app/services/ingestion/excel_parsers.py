# -*- coding: utf-8 -*-
"""
orbita_ingest.excel_parsers
============================
Parsers para os ficheiros .xlsx:
  - aurea_carteira          (Carteira (XX).xlsx)            -> portfolio_snapshots
  - aurea_destaques         (Bolsa-Destaques (XX).xlsx)     -> market_snapshots
  - aurea_ordens / bodiva_ordens (Ordens_Disponiveis*.xlsx) -> order_book_snapshots
  - bodiva_resumo_mercados  (Resumo_dos_Mercados*.xlsx)     -> market_snapshots

Biblioteca: openpyxl (data_only=True para ler valores calculados, não fórmulas).

IMPORTANTE - correcções face aos documentos antigos:
  * portfolio_snapshots usa os nomes REAIS confirmados por PRAGMA (S4):
    quantity_total, par_value_unit, quote_price, current_value, acquisition_value,
    current_value_aoa, unrealized_pnl, daily_variation_*, weight_pct (NÃO 'quantity',
    'avg_price', 'current_price').
  * Aurea Destaques: a coluna "% Var. diária" JÁ vem em pontos percentuais
    (-0,952 = -0,95%). NÃO multiplicar por 100 (o MASTER multiplicava - bug).
  * BODIVA Resumo: a coluna "Variação (%)" vem como FRACÇÃO (-0,0095 = -0,95%) =>
    multiplicar por 100.
"""
from __future__ import annotations

import os
from collections import defaultdict
from datetime import date, datetime
from typing import Any, List, Optional

import openpyxl

from .common import (
    DBRow,
    IngestResult,
    infer_class_from_aurea,
    map_typology,
    normalize_ticker,
    parse_iso_datetime,
    parse_pt_date,
    timestamp_from_filename,
    to_int,
    to_number,
)


def _open_ws(path: str, sheet_name: Optional[str] = None):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.active


def _cell(ws, r: int, c: int) -> Any:
    return ws.cell(row=r, column=c).value


# =========================================================================== #
# 1. AUREA CARTEIRA  ->  portfolio_snapshots
# =========================================================================== #
def parse_aurea_carteira(
    path: str,
    file_hash: str,
    user_id: Optional[int],
    portfolio_id: Optional[int] = None,
    snapshot_date: Optional[date] = None,
) -> IngestResult:
    started = datetime.now()
    res = IngestResult("aurea_carteira", file_hash, snapshot_date, user_id)

    if snapshot_date is None:
        # Carteira__14_.xlsx não traz data fiável -> hoje, com aviso.
        snapshot_date = date.today()
        res.warnings.append(
            "snapshot_date não detectável no nome do ficheiro; usado date.today(). "
            "Confirmar com o utilizador no upload."
        )
    res.snapshot_date = snapshot_date

    ws = _open_ws(path)
    rows: List[DBRow] = []
    holdings_tmp = []  # para calcular weight_pct num 2.º passo

    # Linhas 1-2 = cabeçalhos compostos; dados a partir da linha 3.
    for r in range(3, (ws.max_row or 0) + 1):
        ticker = normalize_ticker(_cell(ws, r, 2))
        if not ticker:           # ticker vazio -> fim da tabela / linha de totais
            continue
        title = str(_cell(ws, r, 1) or "").strip()
        market = str(_cell(ws, r, 3) or "").strip()

        qty_total = to_number(_cell(ws, r, 4))
        qty_avail = to_number(_cell(ws, r, 5))
        par_unit = to_number(_cell(ws, r, 6))
        quote = to_number(_cell(ws, r, 7))
        currency = (str(_cell(ws, r, 8) or "AOA").strip() or "AOA")
        cur_value_aoa = to_number(_cell(ws, r, 9))
        acq_value = to_number(_cell(ws, r, 10))

        if cur_value_aoa is None and quote is None:
            res.rows_skipped += 1
            res.warnings.append("Linha %d ignorada: sem cotação nem valor." % r)
            continue

        instrument_class = infer_class_from_aurea(market, title)
        pnl = (cur_value_aoa - acq_value) if (cur_value_aoa is not None and acq_value is not None) else None
        pnl_pct = (round(100.0 * pnl / acq_value, 4)) if (pnl is not None and acq_value) else None

        values = {
            "user_id": user_id,
            "portfolio_id": portfolio_id,
            "broker": "AUREA",
            "ticker": ticker,
            "title": title,
            "market": market,
            "instrument_class": instrument_class,
            "snapshot_date": snapshot_date.isoformat(),
            "quantity_total": qty_total,
            "quantity_available": qty_avail,
            "par_value_unit": par_unit,
            "quote_price": quote,
            "currency": currency,
            "current_value": cur_value_aoa,        # Aurea reporta em AOA
            "acquisition_value": acq_value,
            "current_value_aoa": cur_value_aoa,
            "unrealized_pnl": pnl,
            "unrealized_pnl_pct": pnl_pct,
            "daily_variation_aoa": None,
            "daily_variation_pct": None,
            "weight_pct": None,                    # preenchido no 2.º passo
        }
        holdings_tmp.append(values)
        res.rows_processed += 1

        # garantir entrada (stub) em bond_master
        rows.append(DBRow(
            "bond_master",
            {
                "ticker": ticker, "title": title, "isin": None,
                "instrument_class": instrument_class, "currency": currency,
                "par_value": par_unit, "data_source": "aurea_carteira",
            },
            conflict_keys=["ticker"],
        ))

    # 2.º passo: weight_pct
    total = sum((h["current_value_aoa"] or 0.0) for h in holdings_tmp)
    for h in holdings_tmp:
        if total > 0 and h["current_value_aoa"] is not None:
            h["weight_pct"] = round(100.0 * h["current_value_aoa"] / total, 4)
        rows.append(DBRow(
            "portfolio_snapshots", h,
            conflict_keys=["broker", "ticker", "snapshot_date"],
        ))

    res.rows = rows
    res.summary = "%d holdings, total %.0f AOA" % (len(holdings_tmp), total)
    if len(holdings_tmp) < 1:
        res.errors.append("Nenhum holding extraído - estrutura inesperada.")
    res.duration_seconds = (datetime.now() - started).total_seconds()
    return res


# =========================================================================== #
# 2. AUREA DESTAQUES  ->  market_snapshots
# =========================================================================== #
def parse_aurea_destaques(path: str, file_hash: str, user_id: Optional[int] = None) -> IngestResult:
    started = datetime.now()
    res = IngestResult("aurea_destaques", file_hash, None, user_id)
    ws = _open_ws(path)
    rows: List[DBRow] = []
    latest_dt: Optional[datetime] = None

    # Cabeçalhos na linha 1; dados a partir da linha 2.
    for r in range(2, (ws.max_row or 0) + 1):
        ticker = normalize_ticker(_cell(ws, r, 2))
        if not ticker:
            continue
        price = to_number(_cell(ws, r, 3))
        currency = str(_cell(ws, r, 5) or "AOA").strip() or "AOA"
        ts = parse_iso_datetime(_cell(ws, r, 6))
        # CORRECÇÃO: já vem em % -> NÃO multiplicar por 100
        var_pct = to_number(_cell(ws, r, 8))
        best_bid = to_number(_cell(ws, r, 9))
        best_ask = to_number(_cell(ws, r, 10))
        volume_trades = to_int(_cell(ws, r, 11))

        snap_dt = ts or datetime.now()
        if latest_dt is None or snap_dt > latest_dt:
            latest_dt = snap_dt

        rows.append(DBRow(
            "market_snapshots",
            {
                "ticker": ticker,
                "snapshot_date": snap_dt.date().isoformat(),
                "snapshot_time": snap_dt.time().strftime("%H:%M:%S"),
                "source": "aurea_destaques",
                "instrument_class": None,
                "price": price,
                "currency": currency,
                "var_daily_pct": var_pct,
                "best_bid": best_bid,
                "best_ask": best_ask,
                "n_trades_day": None,
                "volume_qty": None,
                "volume_aoa": None,
                "volume_trades": volume_trades,
            },
            conflict_keys=["ticker", "snapshot_date", "snapshot_time", "source"],
        ))
        res.rows_processed += 1

    res.snapshot_date = latest_dt.date() if latest_dt else date.today()
    res.rows = rows
    res.summary = "%d instrumentos (destaques Aurea)" % res.rows_processed
    if res.rows_processed < 1:
        res.errors.append("Nenhuma linha de mercado extraída.")
    res.duration_seconds = (datetime.now() - started).total_seconds()
    return res


# =========================================================================== #
# 3. BODIVA RESUMO DOS MERCADOS  ->  market_snapshots
# =========================================================================== #
def parse_bodiva_resumo(path: str, file_hash: str, user_id: Optional[int] = None) -> IngestResult:
    started = datetime.now()
    res = IngestResult("bodiva_resumo_mercados", file_hash, None, user_id)

    snap_dt = timestamp_from_filename(path) or datetime.now()
    if timestamp_from_filename(path) is None:
        res.warnings.append("Timestamp não extraído do nome; usado datetime.now().")
    res.snapshot_date = snap_dt.date()

    ws = _open_ws(path, "Resumo dos Mercados")
    rows: List[DBRow] = []

    # Linha 1 cosmética, linha 2 cabeçalhos; dados a partir da linha 3.
    for r in range(3, (ws.max_row or 0) + 1):
        ticker = normalize_ticker(_cell(ws, r, 1))
        if not ticker:
            continue
        typology = str(_cell(ws, r, 2) or "").strip()
        price = to_number(_cell(ws, r, 3))
        var_frac = to_number(_cell(ws, r, 4))     # FRACÇÃO -> ×100
        n_trades = to_int(_cell(ws, r, 5))
        quantity = to_number(_cell(ws, r, 6))
        montante = to_number(_cell(ws, r, 7))

        rows.append(DBRow(
            "market_snapshots",
            {
                "ticker": ticker,
                "snapshot_date": snap_dt.date().isoformat(),
                "snapshot_time": snap_dt.time().strftime("%H:%M:%S"),
                "source": "bodiva_resumo",
                "instrument_class": map_typology(typology),
                "price": price,
                "currency": "AOA",
                "var_daily_pct": (var_frac * 100.0) if var_frac is not None else None,
                "best_bid": None,
                "best_ask": None,
                "n_trades_day": n_trades,
                "volume_qty": quantity,
                "volume_aoa": montante,
                "volume_trades": None,
            },
            conflict_keys=["ticker", "snapshot_date", "snapshot_time", "source"],
        ))
        res.rows_processed += 1

    res.rows = rows
    res.summary = "%d instrumentos (resumo BODIVA) @ %s" % (
        res.rows_processed, snap_dt.strftime("%d/%m/%Y %H:%M"))
    if res.rows_processed < 1:
        res.errors.append("Nenhuma linha de resumo extraída.")
    res.duration_seconds = (datetime.now() - started).total_seconds()
    return res


# =========================================================================== #
# 4. ORDENS DISPONÍVEIS (Aurea / BODIVA)  ->  order_book_snapshots
# =========================================================================== #
def parse_ordens_disponiveis(
    path: str,
    file_hash: str,
    source: str = "aurea_ordens",
    user_id: Optional[int] = None,
) -> IngestResult:
    """
    Livro de ordens em formato 'long': cada ticker tem várias linhas (níveis).
    Gera, por nível, uma linha BID (compra) e/ou uma linha ASK (venda) em
    order_book_snapshots (schema REAL: ticker, snapshot_date, snapshot_time,
    side, level, quantity, price, yield_pct, last_quote).

    NB: a tabela real NÃO tem coluna 'source'. A proveniência (Aurea vs BODIVA)
    não é persistida aqui - ver recomendação na documentação (adicionar 'source').
    """
    started = datetime.now()
    res = IngestResult(source, file_hash, None, user_id)

    snap_dt = timestamp_from_filename(path) or datetime.now()
    if timestamp_from_filename(path) is None:
        res.warnings.append("Timestamp não extraído do nome; usado datetime.now().")
    res.snapshot_date = snap_dt.date()
    sdate = snap_dt.date().isoformat()
    stime = snap_dt.time().strftime("%H:%M:%S")

    ws = _open_ws(path, "Ordens Disponíveis no Mercado")
    rows: List[DBRow] = []
    levels_by_ticker = defaultdict(int)

    # Linhas 1-2 cosméticas/parciais, headers na linha 3; dados a partir da 4.
    for r in range(4, (ws.max_row or 0) + 1):
        ticker = normalize_ticker(_cell(ws, r, 1))
        if not ticker:
            continue
        isin = str(_cell(ws, r, 2) or "").strip() or None
        typology = str(_cell(ws, r, 3) or "").strip()
        coupon = to_number(_cell(ws, r, 5))
        issue_date = parse_pt_date(_cell(ws, r, 6))
        maturity_date = parse_pt_date(_cell(ws, r, 7))
        last_price = to_number(_cell(ws, r, 8))

        buy_qty = to_number(_cell(ws, r, 9))
        buy_price = to_number(_cell(ws, r, 10))
        yield_val = to_number(_cell(ws, r, 11))
        sell_qty = to_number(_cell(ws, r, 12))
        sell_price = to_number(_cell(ws, r, 13))

        levels_by_ticker[ticker] += 1
        level = levels_by_ticker[ticker]

        if buy_qty is not None and buy_price is not None:
            rows.append(DBRow(
                "order_book_snapshots",
                {
                    "ticker": ticker, "snapshot_date": sdate, "snapshot_time": stime,
                    "side": "BID", "level": level, "quantity": buy_qty,
                    "price": buy_price, "yield_pct": yield_val, "last_quote": last_price,
                },
                conflict_keys=["ticker", "snapshot_date", "snapshot_time", "side", "level"],
            ))
        if sell_qty is not None and sell_price is not None:
            rows.append(DBRow(
                "order_book_snapshots",
                {
                    "ticker": ticker, "snapshot_date": sdate, "snapshot_time": stime,
                    "side": "ASK", "level": level, "quantity": sell_qty,
                    "price": sell_price, "yield_pct": None, "last_quote": last_price,
                },
                conflict_keys=["ticker", "snapshot_date", "snapshot_time", "side", "level"],
            ))

        # stub bond_master (auto-seed de tickers do mercado)
        rows.append(DBRow(
            "bond_master",
            {
                "ticker": ticker, "isin": isin,
                "instrument_class": map_typology(typology),
                "typology": typology or None, "coupon_rate": coupon,
                "issue_date": issue_date.isoformat() if issue_date else None,
                "maturity_date": maturity_date.isoformat() if maturity_date else None,
                "data_source": source,
            },
            conflict_keys=["ticker"],
        ))
        res.rows_processed += 1

    res.rows = rows
    res.summary = "%d tickers no livro de ordens (%s)" % (len(levels_by_ticker), source)
    if res.rows_processed < 1:
        res.errors.append("Livro de ordens vazio - estrutura inesperada.")
    res.duration_seconds = (datetime.now() - started).total_seconds()
    return res
