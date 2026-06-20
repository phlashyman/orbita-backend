# -*- coding: utf-8 -*-
"""
smoke_test - validação dos parsers Órbita
=========================================
Três modos:

  1) CONTRA OS TEUS FICHEIROS REAIS (sem base de dados):
         python3 smoke_test.py /caminho/para/pasta_com_ficheiros
     Detecta o tipo e processa cada ficheiro (preview), imprimindo
     tipo detectado, score, data, nº de linhas e avisos/erros.

  2) AUTO-TESTE SINTÉTICO (sem base de dados) - corre sem argumentos:
         python3 smoke_test.py
     Gera ficheiros sintéticos dos 8 parsers e verifica detecção,
     parsing e as duas correcções de variação. Não precisa de DB.

  3) IDEMPOTÊNCIA EM POSTGRESQL REAL (opcional) - define a env var:
         ORBITA_PG_DSN="postgresql://user:pass@localhost/orbita" python3 smoke_test.py
     Aplica uma carteira sintética duas vezes via asyncpg e confirma
     success -> skipped. (Requer asyncpg e a schema_postgresql.sql aplicada.)

Requisitos: openpyxl, pdfplumber. Para gerar PDFs sintéticos: reportlab.
"""
import os
import sys
import tempfile

from app.services.ingestion.service import parse_file, detect_only, ForbiddenFile, UndetectedFile  # noqa: E402

PASS, FAIL = [], []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print("  [%s] %s %s" % ("OK " if cond else "XX ", name, ("- " + detail) if detail else ""))


def _rows(result, table):
    return [r for r in result.rows if r.table == table]


# --------------------------------------------------------------------------- #
# MODO 1 - ficheiros reais
# --------------------------------------------------------------------------- #
def run_real(folder):
    print("A processar ficheiros reais em:", folder)
    exts = (".xlsx", ".xls", ".pdf")
    files = [os.path.join(folder, f) for f in sorted(os.listdir(folder))
             if f.lower().endswith(exts)]
    if not files:
        print("  (nenhum .xlsx/.pdf encontrado)"); return
    for path in files:
        det = detect_only(path)
        print("\n• %s" % os.path.basename(path))
        print("    detectado: %s  (score %.2f, admin_only=%s)"
              % (det["detected"], det["score"], det["admin_only"]))
        try:
            res = parse_file(path, user_id=3, is_admin=True)  # admin p/ não bloquear institucionais
        except ForbiddenFile as e:
            print("    [papel] %s" % e); continue
        except UndetectedFile as e:
            print("    [XX] não detectado: %s" % e); continue
        print("    parser: %s | data: %s | linhas: %d | processadas: %d"
              % (res.parser_name,
                 res.snapshot_date.isoformat() if res.snapshot_date else "-",
                 len(res.rows), res.rows_processed))
        print("    resumo: %s" % res.summary)
        for w in res.warnings:
            print("    aviso: %s" % w)
        for e in res.errors:
            print("    ERRO: %s" % e)


# --------------------------------------------------------------------------- #
# MODO 2 - sintético (sem DB)
# --------------------------------------------------------------------------- #
def _make_synth(tmp):
    import openpyxl
    paths = {}

    # Aurea Carteira
    p = os.path.join(tmp, "Carteira__14_.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Carteira"
    ws.append(["Título", "Ticker", "Mercado", "Quantidade", "Q. Disp.",
               "Valor Nominal", "Cotação", "Moeda", "Valor AOA", "Valor Aquisição"])
    ws.append([None] * 10)
    ws.append(["BAI", "BAI", "BODIVA AÇÕES", 1000, 1000, 1000, 1450.0, "AOA", 1450000.0, 1200000.0])
    ws.append(["UGD Tesouro", "UGD2027", "BODIVA OBRIGAÇÕES", 500, 500, 100000, 98.5, "AOA", 49250000.0, 48000000.0])
    wb.save(p); paths["carteira"] = p

    # Aurea Destaques (% var. diária JÁ em %)
    p = os.path.join(tmp, "Bolsa-Destaques.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Destaques"
    ws.append(["Título", "Ticker", "Últ. Cotação", "Var", "Moeda",
               "Data/Hora", "x", "% Var. diária", "Compra", "Venda", "Volume"])
    ws.append(["BAI", "BAI", 1450.0, None, "AOA", "2026-05-14T10:38:4", None, -0.952, 1448.0, 1452.0, 37])
    wb.save(p); paths["destaques"] = p

    # Ordens
    p = os.path.join(tmp, "Ordens_Disponiveis_14-05-2026_10-37-21.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Ordens Disponíveis no Mercado"
    ws.append(["Ordens Disponíveis no Mercado"]); ws.append([None])
    ws.append(["Ticker", "ISIN", "Tipologia", "x", "Cupão", "Emissão",
               "Maturidade", "Últ", "Q.Compra", "P.Compra", "Yield", "Q.Venda", "P.Venda"])
    ws.append(["UGD2027", "AOUGDOIF25A1", "OT-TV", None, 16.75, "14/05/2025",
               "14/05/2027", 98.5, 1000, 98.40, 17.10, 800, 98.70])
    wb.save(p); paths["ordens"] = p

    # Resumo (variação como FRACÇÃO)
    p = os.path.join(tmp, "Resumo_dos_Mercados_14-05-2026_10-37-21.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Resumo dos Mercados"
    ws.append(["Resumo"])
    ws.append(["Valor Mobiliário", "Tipologia", "Preço", "Variação (%)",
               "N° de Negócios", "Quantidade", "Montante"])
    ws.append(["BAI", "Acções", 1450.0, -0.0095, 12, 5000, 7250000.0])
    wb.save(p); paths["resumo"] = p

    # PDFs (se reportlab existir)
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4

        def pdf(name, lines):
            pp = os.path.join(tmp, name)
            c = canvas.Canvas(pp, pagesize=A4); c.setFont("Helvetica", 10); y = 800
            for ln in lines:
                c.drawString(40, y, ln); y -= 16
            c.save(); return pp

        paths["standard"] = pdf("A Minha Carteira.pdf", [
            "A MINHA CARTEIRA", "PRODUTO QUANTIDADE PRECO VALOR ACTUAL GANHOS NAO REALIZADOS",
            "STANDARD TESOURARIA FEIVMA 1.234,5678 1.045,20 1.290.000,00 AOA 45.000,00 AOA"])
        paths["ficha"] = pdf("Ficha_Titulos.pdf", [
            "FICHA TECNICA - TITULOS PUBLICOS (OBRIGACOES DO TESOURO)",
            "ISIN: AOUGDOIF25A1", "Codigo de Negociacao: OTTV2025",
            "Emitente: Unidade de Gestao da Divida Publica", "Tipologia: OT-TV",
            "Moeda: AOA", "Valor Nominal: 100.000,00", "Taxa de Cupao: 16,75",
            "Frequencia: Semestral", "Data de Emissao: 14/05/2025",
            "Data de Vencimento: 14/05/2027"])
        paths["boletim"] = pdf("BoletimDiario_20260514.pdf", [
            "BODIVA - BOLSA DE DIVIDA E VALORES DE ANGOLA", "BOLETIM DE MERCADO",
            "Boletim de Mercado Nº 1234", "14 de Maio de 2026",
            "OTTV2025 OT-TV 14/05/2025 14/05/2027 16,75 17,10 5 1000 100,0 101,0 99,5 100,2 100,5 0,30",
            "Curva de Rendimentos", "3M 13,84% 13,79%", "6M 14,20% 14,11%"])
    except ImportError:
        print("  (reportlab ausente - PDFs sintéticos saltados)")
    return paths


def run_synth():
    tmp = tempfile.mkdtemp(prefix="orbita_smoke_")
    print("Auto-teste sintético em:", tmp)
    p = _make_synth(tmp)

    print("\n== DETECÇÃO ==")
    expect = {
        "carteira": "aurea_carteira", "destaques": "aurea_destaques",
        "ordens": "ordens_disponiveis", "resumo": "bodiva_resumo",
        "standard": "standard_carteira", "ficha": "ficha_tecnica",
        "boletim": "bodiva_boletim",
    }
    for k, exp in expect.items():
        if k not in p:
            continue
        det = detect_only(p[k])
        check("detect:%s" % exp, det["detected"] == exp, "obtido=%s" % det["detected"])

    print("\n== PARSING (preview, sem DB) ==")
    r = parse_file(p["carteira"], 3, portfolio_id=1)
    ps = _rows(r, "portfolio_snapshots")
    check("carteira: 2 holdings", len(ps) == 2)
    wsum = sum((row.values.get("weight_pct") or 0) for row in ps)
    check("carteira: weight ~100", abs(wsum - 100.0) < 0.5, "soma=%.3f" % wsum)

    r = parse_file(p["destaques"], 3)
    v = _rows(r, "market_snapshots")[0].values["var_daily_pct"]
    check("destaques: var NAO x100 (-0,952)", abs(v - (-0.952)) < 1e-6, "v=%s" % v)

    r = parse_file(p["resumo"], 1, is_admin=True)
    v = _rows(r, "market_snapshots")[0].values["var_daily_pct"]
    check("resumo: var x100 (-0,95)", abs(v - (-0.95)) < 1e-6, "v=%s" % v)

    r = parse_file(p["ordens"], 1, is_admin=True, source_hint="bodiva")
    ob = _rows(r, "order_book_snapshots")
    sides = {row.values["side"] for row in ob}
    check("ordens: BID e ASK", sides == {"BID", "ASK"}, str(sorted(sides)))

    # regra de papéis
    try:
        parse_file(p["resumo"], 3, is_admin=False)
        check("resumo bloqueado p/ user", False, "não levantou ForbiddenFile")
    except ForbiddenFile:
        check("resumo bloqueado p/ user", True)

    if "standard" in p:
        r = parse_file(p["standard"], 3)
        sp = _rows(r, "portfolio_snapshots")
        check("standard: FEIVMA", len(sp) == 1 and abs(sp[0].values["quote_price"] - 1045.20) < 1e-6)
    if "ficha" in p:
        r = parse_file(p["ficha"], 1, is_admin=True)
        bm = _rows(r, "bond_master")
        check("ficha: BOND_GOV", len(bm) == 1 and bm[0].values["instrument_class"] == "BOND_GOV")
    if "boletim" in p:
        r = parse_file(p["boletim"], 1, is_admin=True)
        ms = _rows(r, "market_snapshots")
        yc = _rows(r, "yield_curve_history")
        check("boletim: negócio->market", len(ms) >= 1 and abs(ms[0].values["price"] - 100.5) < 1e-6)
        check("boletim: 2 pontos de curva", len(yc) == 2)

    print("\n== RESUMO ==  %d OK, %d FALHAS" % (len(PASS), len(FAIL)))
    if FAIL:
        print("  FALHAS:", FAIL); sys.exit(1)
    print("  TODOS OS TESTES PASSARAM.")


# --------------------------------------------------------------------------- #
# MODO 3 - idempotência em PostgreSQL real (opcional)
# --------------------------------------------------------------------------- #
def run_pg(dsn):
    import asyncio
    import openpyxl
    try:
        import asyncpg
    except ImportError:
        print("asyncpg não instalado - modo PostgreSQL ignorado."); return
    from persistence_async import apply_ingest_result_async

    tmp = tempfile.mkdtemp(prefix="orbita_pg_")
    p = os.path.join(tmp, "Carteira__test_idem_.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Carteira"
    ws.append(["Título", "Ticker", "Mercado", "Quantidade", "Q. Disp.",
               "Valor Nominal", "Cotação", "Moeda", "Valor AOA", "Valor Aquisição"])
    ws.append([None] * 10)
    ws.append(["TESTE", "ZZTEST", "BODIVA AÇÕES", 1, 1, 1, 1.0, "AOA", 1.0, 1.0])
    wb.save(p)

    async def go():
        conn = await asyncpg.connect(dsn)
        try:
            from datetime import date
            r1 = parse_file(p, user_id=999, snapshot_date=date(2000, 1, 1))
            out1 = await apply_ingest_result_async(conn, r1)
            r2 = parse_file(p, user_id=999, snapshot_date=date(2000, 1, 1))
            out2 = await apply_ingest_result_async(conn, r2)
            check("PG: 1.ª importação = success", out1["status"] == "success", str(out1))
            check("PG: 2.ª importação = skipped", out2["status"] == "skipped", str(out2))
            # limpeza
            await conn.execute("DELETE FROM portfolio_snapshots WHERE ticker='ZZTEST'")
            await conn.execute("DELETE FROM bond_master WHERE ticker='ZZTEST'")
            await conn.execute("DELETE FROM import_log WHERE file_hash=$1", r1.file_hash)
        finally:
            await conn.close()

    print("Modo PostgreSQL real:", dsn.split("@")[-1])
    asyncio.run(go())
    print("\n== RESUMO PG ==  %d OK, %d FALHAS" % (len(PASS), len(FAIL)))
    if FAIL:
        sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) > 1 and os.path.isdir(sys.argv[1]):
        run_real(sys.argv[1])
    elif os.environ.get("ORBITA_PG_DSN"):
        run_pg(os.environ["ORBITA_PG_DSN"])
    else:
        run_synth()
