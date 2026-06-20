"""
Router de integracao orbita_ingest para o backend Orbita.
Adaptado para FastAPI async + PostgreSQL + JWT auth.

Endpoints:
  POST /portfolio/uploads/{id}/preview  - processa ficheiro sem gravar na DB
  POST /portfolio/uploads/{id}/process  - processa e aplica na DB (idempotente)
"""
import hashlib
import time
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.security import get_current_user
from app.database import get_db
from app.models.user import User
from app.models.broker_file_upload import BrokerFileUpload as Upload
from app.schemas.broker_file import BrokerFileProcessResult

from .detect import detect
from .excel_parsers import parse_aurea_carteira
from .pdf_parsers import parse_standard_carteira

router = APIRouter(prefix="/portfolio/uploads", tags=["ingestion"])

# Ficheiros que apenas ADMIN pode carregar
ADMIN_ONLY = {"bodiva_resumo", "bodiva_boletim", "bodiva_relatorio"}


async def _read_upload_file(upload: Upload) -> bytes:
    """Le o ficheiro do local ou S3."""
    if upload.s3_key.startswith("local://"):
        local_path = upload.s3_key.replace("local://", "")
        return Path(local_path).read_bytes()
    else:
        from app.utils.s3 import get_s3_client
        from app.config import settings
        s3 = get_s3_client()
        response = s3.get_object(Bucket=settings.s3_bucket, Key=upload.s3_key)
        return response["Body"].read()


def _detect_parser(upload: Upload) -> str:
    """Detecta qual parser usar baseado no nome e conteudo."""
    filename = upload.original_filename or upload.stored_filename
    file_type = upload.file_type or ""
    return detect(filename, file_type)


def _file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


@router.post("/{upload_id}/preview", response_model=BrokerFileProcessResult)
async def preview_upload(
    upload_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Preview do processamento - extrai posicoes SEM gravar na DB.
    Util para verificar o que vai ser importado antes de aplicar.
    """
    upload = await db.get(Upload, upload_id)
    if not upload or upload.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Upload not found")

    file_bytes = await _read_upload_file(upload)
    parser_key = _detect_parser(upload)

    try:
        if parser_key == "aurea_carteira":
            from openpyxl import load_workbook
            import io
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            result = parse_aurea_carteira(wb, file_hash=_file_hash(file_bytes))
        elif parser_key == "standard_carteira":
            import pdfplumber
            import io
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                result = parse_standard_carteira(pdf, file_hash=_file_hash(file_bytes))
        else:
            return BrokerFileProcessResult(
                upload_id=upload.id,
                status="UNSUPPORTED",
                file_type=upload.file_type,
                positions_found=0,
                positions=[],
                error_message=f"Parser '{parser_key}' ainda nao integrado. Preview disponivel para aurea_carteira e standard_carteira.",
            )

        # Converte DBRows para positions (formato do nosso schema)
        positions = []
        for row in result.rows:
            positions.append({
                "ticker": row.values.get("ticker", ""),
                "name": row.values.get("name", None),
                "quantity": row.values.get("quantity_total", None),
                "price": row.values.get("quote_price", None),
                "value": str(row.values.get("current_value_aoa", None)),
                "instrument_type": row.values.get("instrument_class", ""),
                "raw_data": row.values,
            })

        return BrokerFileProcessResult(
            upload_id=upload.id,
            status="PREVIEW",
            file_type=upload.file_type,
            positions_found=len(positions),
            positions=positions,
            error_message=None,
        )

    except Exception as exc:
        return BrokerFileProcessResult(
            upload_id=upload.id,
            status="ERROR",
            file_type=upload.file_type,
            positions_found=0,
            positions=[],
            error_message=f"Preview failed: {exc}",
        )


@router.post("/{upload_id}/process", response_model=BrokerFileProcessResult)
async def process_upload(
    upload_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Processa o ficheiro e aplica na DB (idempotente).
    Se ja foi importado (mesmo hash), retorna 'skipped'.
    """
    upload = await db.get(Upload, upload_id)
    if not upload or upload.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Upload not found")

    file_bytes = await _read_upload_file(upload)
    parser_key = _detect_parser(upload)

    # Verificar regra de papeis
    if parser_key in ADMIN_ONLY and not current_user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Apenas administradores podem carregar este tipo de ficheiro",
        )

    try:
        if parser_key == "aurea_carteira":
            from openpyxl import load_workbook
            import io
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            result = parse_aurea_carteira(wb, file_hash=_file_hash(file_bytes))
        elif parser_key == "standard_carteira":
            import pdfplumber
            import io
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                result = parse_standard_carteira(pdf, file_hash=_file_hash(file_bytes))
        else:
            upload.status = "UNSUPPORTED"
            upload.error_message = f"Parser '{parser_key}' ainda nao integrado"
            await db.flush()
            return BrokerFileProcessResult(
                upload_id=upload.id,
                status="UNSUPPORTED",
                file_type=upload.file_type,
                positions_found=0,
                positions=[],
                error_message=f"Parser '{parser_key}' ainda nao integrado.",
            )

        # Verifica se ja foi importado (idempotencia nivel 1)
        from sqlalchemy import select, text
        hash_check = await db.execute(
            text("SELECT COUNT(*) FROM import_log WHERE file_hash = :h AND status = 'success'"),
            {"h": result.file_hash}
        )
        if hash_check.scalar() > 0:
            upload.status = "SKIPPED"
            upload.error_message = "Ficheiro ja importado (hash identico)"
            await db.flush()
            return BrokerFileProcessResult(
                upload_id=upload.id,
                status="SKIPPED",
                file_type=upload.file_type,
                positions_found=0,
                positions=[],
                error_message="Ficheiro ja importado anteriormente.",
            )

        # Aplica na DB (idempotencia nivel 2: DELETE + INSERT)
        applied = 0
        warnings_list = list(result.warnings)
        
        for row in result.rows:
            try:
                # DELETE por chave natural
                conflict = {k: row.values[k] for k in row.conflict_keys if k in row.values}
                if conflict:
                    where_clause = " AND ".join(f"{k} = :{k}" for k in conflict)
                    await db.execute(text(f"DELETE FROM {row.table} WHERE {where_clause}"), conflict)
                
                # INSERT
                cols = list(row.values.keys())
                placeholders = ", ".join(f":{c}" for c in cols)
                sql = f"INSERT INTO {row.table} ({', '.join(cols)}) VALUES ({placeholders})"
                await db.execute(text(sql), row.values)
                applied += 1
            except Exception as row_exc:
                warnings_list.append(f"Linha ignorada: {row_exc}")

        # Log da importacao
        await db.execute(text("""
            INSERT INTO import_log (user_id, parser_name, file_hash, snapshot_date, status,
                                   rows_processed, rows_applied, rows_skipped, warnings, errors)
            VALUES (:user_id, :parser_name, :file_hash, :snapshot_date, 'success',
                    :rows_processed, :rows_applied, :rows_skipped, :warnings, :errors)
        """), {
            "user_id": str(current_user.id),
            "parser_name": result.parser_name,
            "file_hash": result.file_hash,
            "snapshot_date": result.snapshot_date,
            "rows_processed": result.rows_processed,
            "rows_applied": applied,
            "rows_skipped": result.rows_skipped,
            "warnings": "; ".join(warnings_list) if warnings_list else None,
            "errors": "; ".join(result.errors) if result.errors else None,
        })

        upload.status = "PROCESSED"
        upload.error_message = None
        await db.commit()

        # Converte para formato de resposta
        positions = []
        for row in result.rows:
            positions.append({
                "ticker": row.values.get("ticker", ""),
                "name": row.values.get("name", None),
                "quantity": row.values.get("quantity_total", None),
                "price": row.values.get("quote_price", None),
                "value": str(row.values.get("current_value_aoa", None)),
                "instrument_type": row.values.get("instrument_class", ""),
                "raw_data": row.values,
            })

        return BrokerFileProcessResult(
            upload_id=upload.id,
            status="PROCESSED",
            file_type=upload.file_type,
            positions_found=len(positions),
            positions=positions,
            error_message="; ".join(warnings_list) if warnings_list else None,
        )

    except Exception as exc:
        await db.rollback()
        upload.status = "ERROR"
        upload.error_message = str(exc)
        await db.commit()
        return BrokerFileProcessResult(
            upload_id=upload.id,
            status="ERROR",
            file_type=upload.file_type,
            positions_found=0,
            positions=[],
            error_message=str(exc),
        )
